"""
CSV/Excel customer import — priority 1 of Stage 2.

Four record types are supported, selected by the `dataset` argument:
customers, family, policies, interactions.

The import is two-phase on purpose. `preview()` parses, validates, and
detects duplicates without writing anything; `commit()` re-runs the same
parse and then writes. The advisor sees exactly what will happen before
it happens, and a file that fails validation never partially lands.

Column names are matched case-insensitively and accept common aliases,
because real advisor exports do not agree on headers.
"""
from __future__ import annotations

import csv
import io
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend_v3.integrations.models import (
    NormalizedContact,
    NormalizedCustomer,
    NormalizedInteraction,
    NormalizedPolicy,
    Provenance,
)

SOURCE_SYSTEM = "csv"

DATASETS = ("customers", "family", "policies", "interactions")

# dataset -> (required logical fields, {logical field: accepted headers})
SCHEMAS: dict[str, dict[str, Any]] = {
    "customers": {
        "required": ("external_id", "full_name"),
        "aliases": {
            "external_id": ("external_id", "customer_id", "id", "crm_id", "reference"),
            "full_name": ("full_name", "name", "customer_name", "full name"),
            "email": ("email", "email_address", "e-mail"),
            "phone": ("phone", "phone_number", "mobile", "contact_number"),
            "life_stage": ("life_stage", "lifestage", "segment", "life stage"),
            "advisor_name": ("advisor_name", "advisor", "agent", "adviser"),
        },
    },
    "family": {
        "required": ("customer_external_id", "full_name", "relationship"),
        "aliases": {
            "customer_external_id": ("customer_external_id", "customer_id", "external_id", "id"),
            "full_name": ("full_name", "name", "member_name"),
            "relationship": ("relationship", "relation", "type"),
            "email": ("email", "email_address"),
            "phone": ("phone", "mobile", "phone_number"),
        },
    },
    "policies": {
        "required": ("customer_external_id", "policy_id", "product_name"),
        "aliases": {
            "customer_external_id": ("customer_external_id", "customer_id", "external_id", "id"),
            "policy_id": ("policy_id", "policy_number", "policy"),
            "product_name": ("product_name", "product", "plan", "plan_name"),
            "line_of_business": ("line_of_business", "lob", "category", "line"),
            "annual_premium": ("annual_premium", "premium", "annual premium"),
            "policy_status": ("policy_status", "status"),
        },
    },
    "interactions": {
        "required": ("customer_external_id", "occurred_at", "summary"),
        "aliases": {
            "customer_external_id": ("customer_external_id", "customer_id", "external_id", "id"),
            "occurred_at": ("occurred_at", "date", "interaction_date", "when"),
            "summary": ("summary", "notes", "description", "subject"),
            "interaction_type": ("interaction_type", "type", "channel"),
            "body": ("body", "transcript", "detail", "content"),
        },
    },
}


@dataclass
class RowError:
    row: int
    message: str
    values: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParseResult:
    dataset: str
    valid_rows: list[dict[str, Any]] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    duplicates_in_file: list[RowError] = field(default_factory=list)
    existing_in_system: list[dict[str, Any]] = field(default_factory=list)
    headers: list[str] = field(default_factory=list)


class ImportError_(ValueError):
    """Raised for a file-level problem — unreadable, empty, wrong columns."""


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportError_("File could not be decoded as text. Save it as UTF-8 CSV and try again.")


def _read_rows(content: bytes, filename: str) -> list[dict[str, str]]:
    """Read CSV directly; convert .xlsx via openpyxl when available."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError_(
                "Excel support needs the openpyxl package. Export the sheet as CSV, "
                "or install openpyxl on the backend."
            ) from None
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ImportError_("The spreadsheet is empty.")
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {headers[i]: ("" if cell is None else str(cell)) for i, cell in enumerate(row) if i < len(headers)}
            for row in rows[1:]
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

    text = _decode(content)
    if not text.strip():
        raise ImportError_("The file is empty.")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.DictReader(io.StringIO(text), dialect=dialect)]


def _build_header_map(dataset: str, headers: list[str]) -> dict[str, str]:
    """Map logical field -> actual header present in the file."""
    aliases = SCHEMAS[dataset]["aliases"]
    normalized = {(h or "").strip().lower().replace("-", "_").replace(" ", "_"): h for h in headers}
    mapping: dict[str, str] = {}
    for logical, accepted in aliases.items():
        for candidate in accepted:
            key = candidate.lower().replace("-", "_").replace(" ", "_")
            if key in normalized:
                mapping[logical] = normalized[key]
                break
    return mapping


def _value(row: dict[str, str], mapping: dict[str, str], logical: str) -> str:
    header = mapping.get(logical)
    if not header:
        return ""
    return (row.get(header) or "").strip()


def _parse_premium(raw: str) -> float | None:
    if not raw:
        return None
    cleaned = raw.replace(",", "").replace("S$", "").replace("$", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _normalize_date(raw: str) -> str | None:
    """Accept the date formats advisor exports actually contain."""
    from datetime import datetime

    raw = (raw or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw[:19] if " " in raw else raw, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return None


def parse(content: bytes, filename: str, dataset: str) -> ParseResult:
    """Parse + validate + detect duplicates. Writes nothing."""
    if dataset not in DATASETS:
        raise ImportError_(f"Unknown dataset '{dataset}'. Expected one of: {', '.join(DATASETS)}.")

    rows = _read_rows(content, filename)
    if not rows:
        raise ImportError_("No data rows found beneath the header row.")

    headers = list(rows[0].keys())
    mapping = _build_header_map(dataset, headers)
    required = SCHEMAS[dataset]["required"]
    missing = [f for f in required if f not in mapping]
    if missing:
        raise ImportError_(
            f"Missing required column(s) for '{dataset}': {', '.join(missing)}. "
            f"Columns found: {', '.join(h for h in headers if h)}."
        )

    result = ParseResult(dataset=dataset, headers=headers)
    seen_keys: dict[str, int] = {}

    for index, row in enumerate(rows, start=2):  # row 1 is the header
        values = {logical: _value(row, mapping, logical) for logical in mapping}

        blanks = [f for f in required if not values.get(f)]
        if blanks:
            result.errors.append(RowError(index, f"Missing required value(s): {', '.join(blanks)}", values))
            continue

        if dataset == "customers":
            key = values["external_id"].lower()
            email = values.get("email", "")
            if email and "@" not in email:
                result.errors.append(RowError(index, f"'{email}' is not a valid email address", values))
                continue
        elif dataset == "policies":
            key = values["policy_id"].lower()
            premium_raw = values.get("annual_premium", "")
            if premium_raw and _parse_premium(premium_raw) is None:
                result.errors.append(RowError(index, f"'{premium_raw}' is not a valid premium amount", values))
                continue
        elif dataset == "family":
            key = f"{values['customer_external_id']}|{values['full_name']}".lower()
        else:
            occurred = _normalize_date(values.get("occurred_at", ""))
            if not occurred:
                result.errors.append(
                    RowError(index, f"'{values.get('occurred_at')}' is not a recognizable date", values)
                )
                continue
            values["occurred_at"] = occurred
            key = f"{values['customer_external_id']}|{occurred}|{values['summary'][:40]}".lower()

        if key in seen_keys:
            result.duplicates_in_file.append(
                RowError(index, f"Duplicate of row {seen_keys[key]} in this file", values)
            )
            continue
        seen_keys[key] = index
        values["_row"] = index
        result.valid_rows.append(values)

    result.existing_in_system = _find_existing(dataset, result.valid_rows)
    return result


def _find_existing(dataset: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Which valid rows already exist — an update rather than an insert."""
    if not rows:
        return []
    from backend_v3.graph_store.neo4j_client import run_query

    if dataset == "customers":
        ids = [r["external_id"] for r in rows]
        found = run_query(
            "MATCH (c:Customer)-[:HAS_IDENTITY]->(i:CustomerIdentity {kind: 'external_id'}) "
            "WHERE i.value_normalized IN $ids "
            "RETURN i.value_normalized AS external_id, c.customer_id AS customer_id, c.name AS name",
            {"ids": [i.lower() for i in ids]},
        )
        return [dict(f, action="update") for f in found]

    if dataset == "policies":
        ids = [r["policy_id"] for r in rows]
        found = run_query(
            "MATCH (p:Policy) WHERE p.policy_id IN $ids "
            "RETURN p.policy_id AS policy_id, p.product_name AS product_name",
            {"ids": ids},
        )
        return [dict(f, action="update") for f in found]

    return []


def to_normalized(result: ParseResult, filename: str) -> dict[str, list[Any]]:
    """Turn validated rows into normalized objects carrying provenance."""
    customers: list[NormalizedCustomer] = []
    contacts: list[NormalizedContact] = []
    policies: list[NormalizedPolicy] = []
    interactions: list[NormalizedInteraction] = []

    for values in result.valid_rows:
        row_number = values.get("_row", 0)
        if result.dataset == "customers":
            provenance = Provenance(
                source_system=SOURCE_SYSTEM,
                source_id=values["external_id"],
                original_reference=f"{filename}:row{row_number}",
            )
            customers.append(
                NormalizedCustomer(
                    external_id=values["external_id"],
                    full_name=values["full_name"],
                    email=values.get("email") or None,
                    phone=values.get("phone") or None,
                    life_stage=values.get("life_stage") or None,
                    advisor_name=values.get("advisor_name") or None,
                    provenance=provenance,
                )
            )
        elif result.dataset == "family":
            provenance = Provenance(
                SOURCE_SYSTEM,
                f"{values['customer_external_id']}:{values['full_name']}",
                f"{filename}:row{row_number}",
            )
            contacts.append(
                NormalizedContact(
                    customer_external_id=values["customer_external_id"],
                    full_name=values["full_name"],
                    relationship=values["relationship"],
                    email=values.get("email") or None,
                    phone=values.get("phone") or None,
                    provenance=provenance,
                )
            )
        elif result.dataset == "policies":
            provenance = Provenance(SOURCE_SYSTEM, values["policy_id"], f"{filename}:row{row_number}")
            policies.append(
                NormalizedPolicy(
                    customer_external_id=values["customer_external_id"],
                    policy_id=values["policy_id"],
                    product_name=values["product_name"],
                    line_of_business=values.get("line_of_business") or None,
                    annual_premium=_parse_premium(values.get("annual_premium", "")),
                    policy_status=values.get("policy_status") or None,
                    provenance=provenance,
                )
            )
        else:
            provenance = Provenance(
                SOURCE_SYSTEM,
                f"{values['customer_external_id']}:{values['occurred_at']}:{row_number}",
                f"{filename}:row{row_number}",
            )
            interactions.append(
                NormalizedInteraction(
                    customer_external_id=values["customer_external_id"],
                    interaction_type=values.get("interaction_type") or "note",
                    occurred_at=values["occurred_at"],
                    summary=values["summary"],
                    body=values.get("body") or None,
                    provenance=provenance,
                )
            )

    return {
        "customers": customers,
        "contacts": contacts,
        "policies": policies,
        "interactions": interactions,
    }


def _summarize(result: ParseResult) -> dict[str, Any]:
    return {
        "dataset": result.dataset,
        "headers": [h for h in result.headers if h],
        "valid_count": len(result.valid_rows),
        "error_count": len(result.errors),
        "duplicate_count": len(result.duplicates_in_file),
        "existing_count": len(result.existing_in_system),
        "new_count": max(len(result.valid_rows) - len(result.existing_in_system), 0),
        "errors": [{"row": e.row, "message": e.message} for e in result.errors[:50]],
        "duplicates": [{"row": e.row, "message": e.message} for e in result.duplicates_in_file[:50]],
        "existing": result.existing_in_system[:50],
        "preview": result.valid_rows[:10],
    }


def preview(content: bytes, filename: str, dataset: str) -> dict[str, Any]:
    """Phase 1: show the advisor what would happen. Writes nothing."""
    result = parse(content, filename, dataset)
    return {**_summarize(result), "committed": False}


def commit(content: bytes, filename: str, dataset: str) -> dict[str, Any]:
    """Phase 2: re-validate, then write valid rows through the pipeline."""
    from backend_v3.integrations.audit import audit
    from backend_v3.integrations.connection_store import mark_connected
    from backend_v3.integrations.pipeline import ingest

    result = parse(content, filename, dataset)
    summary = _summarize(result)

    if not result.valid_rows:
        return {**summary, "committed": False, "reason": "No valid rows to import."}

    normalized = to_normalized(result, filename)
    outcome = ingest(source_system=SOURCE_SYSTEM, **normalized)

    mark_connected("csv", account=filename)
    audit(
        "import.csv",
        subject_id="csv",
        metadata={"dataset": dataset, "file": filename, "counts": outcome["counts"]},
    )

    return {
        **summary,
        "committed": True,
        "imported": outcome["counts"],
        "import_errors": outcome["errors"][:50],
    }
